import requests
import json
import argparse
import sys
import traceback
import openpyxl
import types
import os
import uuid
from ingestutils import sectors
import psycopg2
import csv


DB_ENDPOINT = "localhost"
DB_PORT = 5432
DB_USER = ""
DB_PASSWD = ""
DB_NAME = ""


VISTA_NA_FILL_VALUE = "NA"

class ColumnHeaders:
    SOURCE_IDENTIFIER = "Source identifier "
    SOURCE_LATITUDE = "Source latitude (deg)"
    SOURCE_LONGITUDE = "Source longitude (deg)"
    VISTA_ID = "Vista ID"
    SOURCE_TYPE = "Source type "
    IPCC_SECTOR = "IPCC sector"
    NUMBER_OVERFLIGHTS = "N overflights"
    SOURCE_PERSISTANCE = "Source Persistence (f)"
    CONFIDENCE_IN_PERSISTENCE = "Confidence in persistence "
    Q_SOURCE_KGHR = "Qsource (kg/hr)"
    SIGMA_Q_SOURCE_KGHR = "sQ  (kg/hr)"


ColumnLettersSources = {
    ColumnHeaders.SOURCE_IDENTIFIER : "A",
    ColumnHeaders.SOURCE_LATITUDE : "B",
    ColumnHeaders.SOURCE_LONGITUDE : "C",
    ColumnHeaders.NUMBER_OVERFLIGHTS: "G",
    ColumnHeaders.SOURCE_PERSISTANCE : "H",
    ColumnHeaders.Q_SOURCE_KGHR: "J",
    ColumnHeaders.SIGMA_Q_SOURCE_KGHR: "K",
    ColumnHeaders.VISTA_ID: "D",
    ColumnHeaders.CONFIDENCE_IN_PERSISTENCE: "I"
}


def __check_if_neg9999(value):
    if value == -9999L:
        return None
    else:
        return value


# Sources:
def read_source_record(ws, row_num):
    record_map = {}

    if ws["A%d"%row_num].value == None:
        return None

    # First pass for values
    for key in ColumnLettersSources:
        cell_coord = "%s%d" % (ColumnLettersSources[key], row_num)
        value = ws[cell_coord].value
        record_map[key] = value

    return record_map


def read_sources_sheet(ws):
    sources = []
    for n in range(2, 1000):
        record = read_source_record(ws, n)
        if record is None:
            break
        sources.append(record)
    return sources


def __check_if_divbyzero(value):
    if value == "#DIV/0!" or value == '':
        return None
    else:
        return value


def update_source(source, cur, verbose=False):
    sql = """
        update sources set
            source_latitude_deg = %s,
            source_longitude_deg = %s,
            total_overflights = %s,
            Source_persistence = %s,
            q_source_final = %s,
            q_source_final_sigma = %s,
            vista_id = %s,
            source_location = ST_GeomFromText('POINT(%s %s)', 4326),
            confidence_in_persistence = %s,
            is_active = true
          where 
            source_id = %s;
    """
    cur.execute(sql,
                (
                    source[ColumnHeaders.SOURCE_LATITUDE],
                    source[ColumnHeaders.SOURCE_LONGITUDE],
                    source[ColumnHeaders.NUMBER_OVERFLIGHTS],
                    __check_if_divbyzero(source[ColumnHeaders.SOURCE_PERSISTANCE]),
                    __check_if_divbyzero(source[ColumnHeaders.Q_SOURCE_KGHR]),
                    __check_if_divbyzero(source[ColumnHeaders.SIGMA_Q_SOURCE_KGHR]),
                    source[ColumnHeaders.VISTA_ID] if source[ColumnHeaders.VISTA_ID] != VISTA_NA_FILL_VALUE else None,
                    source[ColumnHeaders.SOURCE_LONGITUDE],
                    source[ColumnHeaders.SOURCE_LATITUDE],
                    source[ColumnHeaders.CONFIDENCE_IN_PERSISTENCE],
                    source[ColumnHeaders.SOURCE_IDENTIFIER]
                ))


def insert_source(source, cur, verbose=False):
    sql = """
        INSERT INTO sources
          (
            source_id,
            source_latitude_deg,
            source_longitude_deg,
            total_overflights,
            Source_persistence,
            q_source_final,
            q_source_final_sigma,
            vista_id,
            source_location,
            confidence_in_persistence,
            is_active
          ) VALUES (
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            ST_GeomFromText('POINT(%s %s)', 4326),
            %s,
            true
          )
    """

    cur.execute(sql,
                (
                    source[ColumnHeaders.SOURCE_IDENTIFIER],
                    source[ColumnHeaders.SOURCE_LATITUDE],
                    source[ColumnHeaders.SOURCE_LONGITUDE],
                    source[ColumnHeaders.NUMBER_OVERFLIGHTS],
                    __check_if_divbyzero(source[ColumnHeaders.SOURCE_PERSISTANCE]),
                    __check_if_divbyzero(source[ColumnHeaders.Q_SOURCE_KGHR]),
                    __check_if_divbyzero(source[ColumnHeaders.SIGMA_Q_SOURCE_KGHR]),
                    source[ColumnHeaders.VISTA_ID] if source[ColumnHeaders.VISTA_ID] != VISTA_NA_FILL_VALUE else None,
                    source[ColumnHeaders.SOURCE_LONGITUDE],
                    source[ColumnHeaders.SOURCE_LATITUDE],
                    source[ColumnHeaders.CONFIDENCE_IN_PERSISTENCE]
                ))



def upload_source(source, cur, test_only=False, verbose=False):

    if verbose:
        print source

    if does_source_exist(source[ColumnHeaders.SOURCE_IDENTIFIER], cur) is True:
        if verbose:
            print("Source '%s' already in database. Updating"%source[ColumnHeaders.SOURCE_IDENTIFIER])
        update_source(source, cur, verbose)
    else:
        if verbose:
            print("Source '%s' not yet in database. Inserting"%source[ColumnHeaders.SOURCE_IDENTIFIER])
        insert_source(source, cur, verbose)


def does_source_exist(source_id, cur):
    cur.execute("select count(source_id) as source_count from sources where source_id=%s",
                (source_id,))
    row = cur.fetchone()
    return row[0] >= 1


def set_all_in_db_inactive(cur):
    sql = "update sources set is_active=false;"
    cur.execute(sql)


def upload_sources(sources, test_only=False, verbose=False):
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER, password=DB_PASSWD, host=DB_ENDPOINT, port=DB_PORT)
    cur = conn.cursor()

    set_all_in_db_inactive(cur)

    for source in sources:
        upload_source(source, cur, test_only, verbose)

    if test_only:
        conn.rollback()
    else:
        conn.commit()
    cur.close()
    conn.close()


def process_workbook(input_file, worksheet_name, test_only=False, verbose=False):
    print "Processing", input_file
    wb2 = openpyxl.load_workbook(input_file)

    ws_sources = wb2[worksheet_name]

    sources = read_sources_sheet(ws_sources)

    upload_sources(sources, test_only=test_only, verbose=verbose)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-d", "--data", help="Input Sources Lists (xlsx)", required=True, type=str, nargs='+')
    parser.add_argument("-e", "--endpoint", help="Target PostgreSQL endpoint", type=str, default="localhost", required=False)
    parser.add_argument("-p", "--port", help="Target PostgreSQL port", type=str, default="8983", required=False)
    parser.add_argument("-s", "--sheet", help="Sources worksheet name", type=str, default="Sheet1", required=False)
    parser.add_argument("-v", "--verbose",
                        help="Extra output",
                        required=False, action="store_true")
    parser.add_argument("-t", "--test",
                        help="Test. Don't actually upload to PostgreSQL.",
                        required=False, action="store_true")

    args = parser.parse_args()
    input_files = args.data
    verbose = args.verbose
    test_only = args.test
    worksheet_name = args.sheet

    for input_file in input_files:
        process_workbook(input_file, worksheet_name, test_only=test_only, verbose=verbose)


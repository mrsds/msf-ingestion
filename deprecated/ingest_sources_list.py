import requests
import json
import argparse
import ingestutils.counts as counts
import sys
import traceback
import openpyxl
import types
import os
import uuid

#SOLR_URL = "http://100.64.114.155:8983"
#SOLR_URL = "http://localhost:8983"

SOLR_URL = "http://%s:%s"

ColumnLettersSources = {
    "Source_identifier" : "A",
    "Source_Latitude" : "B",
    "Source_Longitude" : "C",
    "Area" : "D",
    "Source_type" : "E",
    "Nearest_facility" : "F",
    "Sectors" : "G",
    "Selection_crieria" : "H"
}

ColumnLettersPlumes = {
    "Source_identifier" : "A",
    "Source_Latitude" : "B",
    "Source_Longitude" : "C",
    "Plume_identifier" : lambda rm: "%s_%s_%05d"%(rm["Instrument_abbreviation"], rm["Gas"], int(rm["Number"])),
    "Instrument_abbreviation" : "E",
    "Gas" : "F",
    "Number" : "G",
    "Plume_Latitude" : "H",
    "Plume_Longitude" : "I",
    "Selection_crieria" : "J",
    "Area" : "K",
    "Image_analyst" : "L",
    "Image_analyst_confidence" : "M",
    "Candidate_ID" : "N",
    "Source_type" : "O",
    "Nearest_facility" : "P",
    "Sectors" : "Q",
    "Line_name" : "R",
    "Date_of_detection" : lambda rm: "%s/%s/%s"%(rm["Line_name"][7:9], rm["Line_name"][9:11], rm["Line_name"][3:7]),
    "Time_of_detection" : lambda rm: "%s:%s:%s"%(rm["Line_name"][12:14], rm["Line_name"][14:16], rm["Line_name"][16:18]),
    "Candidate_id" : "V",
    "Source_id" : "W",

    "IME5_500ppmm" : "X",
    "IME10_500ppmm" : "Y",
    "IME20_500ppmm" : "Z",
    "Fetch5_500ppmm" : "AA",
    "Fetch10_500ppmm" : "AB",
    "Fetch20_500ppmm" : "AC",
    "DetId5_500ppmm" : "AD",
    "DetId10_500ppmm" : "AE",
    "DetId20_500ppmm" : "AF",

    "IME5_1000ppmm" : "AH",
    "IME10_1000ppmm" : "AI",
    "IME20_1000ppmm" : "AJ",
    "Fetch5_1000ppmm" : "AK",
    "Fetch10_1000ppmm" : "AL",
    "Fetch20_1000ppmm" : "AM",
    "DetId5_1000ppmm" : "AN",
    "DetId10_1000ppmm" : "AO",
    "DetId20_1000ppmm" : "AP",

    "IME5_1500ppmm" : "AR",
    "IME10_1500ppmm" : "AS",
    "IME20_1500ppmm" : "AT",
    "Fetch5_1500ppmm" : "AU",
    "Fetch10_1500ppmm" : "AV",
    "Fetch20_1500ppmm" : "AW",
    "DetId5_1500ppmm" : "AX",
    "DetId10_1500ppmm" : "AY",
    "DetId20_1500ppmm" : "AZ"
}



def read_plume_record(ws, row_num, last_record):
    record_map = {}

    if ws["A%d"%row_num].value == None:
        return None

    # First pass for values
    for key in ColumnLettersPlumes:
        if isinstance(ColumnLettersPlumes[key], types.StringType) is True:
            cell_coord = "%s%d"%(ColumnLettersPlumes[key], row_num)
            value = ws[cell_coord].value
            record_map[key] = value

    if not isinstance(record_map["Number"], types.IntType):
        if last_record is not None:
            record_map["Number"] = last_record["Number"] + 1
        else:
            record_map["Number"] = 1

            # Second pass for formulae (lambdas in our case, tho)
    for key in ColumnLettersPlumes:
        if isinstance(ColumnLettersPlumes[key], types.LambdaType):
            value = ColumnLettersPlumes[key](record_map)
            record_map[key] = value
    return record_map


def read_plume_sheet(ws):
    last_record = None
    plumes = []

    for n in range(5, 1000):
        record = read_plume_record(ws, n, last_record)

        if record is None:
            break

        plumes.append(record)
        last_record = record

    return plumes



def upload_plume_to_solr(plume,  solr_host="localhost:8983",test_only=False, verbose=False):
    id = str(uuid.uuid4())
    name = plume["Plume_identifier"]
    doc = {
        "id": id,
        "record_type": "sources-plume"
    }

    for key in plume:
        value = plume[key]
        if isinstance(value, types.IntType) or isinstance(value, types.FloatType):
            solr_field = ("%s_f" % key).lower()
        else:
            solr_field = ("%s_s"%key).lower()
        doc[solr_field] = value

    if not test_only:
        print "Uploading %s to Solr" % name
        r = requests.post('%s/solr/MSF/update/json/docs?commit=true' % solr_host, data=json.dumps(doc),
                          headers={"Content-Type": "application/json"})

        if r.status_code != 200:
            print r.status_code, r.text
            raise Exception("Failed to upload document. Solr status: %s, %s" % (r.status_code, r.text))
    else:
        print "Test of %s to Solr" % name


def upload_plumes_to_solr(plumes, solr_host="localhost:8983",test_only=False, verbose=False):
    for plume in plumes:
        upload_plume_to_solr(plume, solr_host=solr_host, test_only=test_only, verbose=verbose)



def upload_source_to_solr(source, solr_host="localhost:8983",test_only=False, verbose=False):
    id = str(uuid.uuid4())
    name = source["Source_identifier"]
    doc = {
        "id": id,
        "record_type": "sources-source"
    }

    for key in source:
        value = source[key]
        if isinstance(value, types.IntType) or isinstance(value, types.FloatType):
            solr_field = ("%s_f" % key).lower()
        else:
            solr_field = ("%s_s"%key).lower()
        doc[solr_field] = value

    if not test_only:
        print "Uploading %s to Solr" % name
        r = requests.post('%s/solr/MSF/update/json/docs?commit=true' % solr_host, data=json.dumps(doc),
                          headers={"Content-Type": "application/json"})

        if r.status_code != 200:
            print r.status_code, r.text
            raise Exception("Failed to upload document. Solr status: %s, %s" % (r.status_code, r.text))
    else:
        print "Test of %s to Solr" % name


def upload_sources_to_solr(sources, solr_host="localhost:8983",test_only=False, verbose=False):
    for source in sources:
        upload_source_to_solr(source, solr_host=solr_host, test_only=test_only, verbose=verbose)


def read_source_record(ws, row_num):
    record_map = {}

    if ws["A%d"%row_num].value == None:
        return None

    # First pass for values
    for key in ColumnLettersSources:
        cell_coord = "%s%d" % (ColumnLettersSources[key], row_num)
        value = ws[cell_coord].value
        record_map[key] = value

    return record_map


def read_sources_sheet(ws):
    sources = []
    for n in range(2, 1000):
        record = read_source_record(ws, n)
        if record is None:
            break
        sources.append(record)
    return sources



def process_workbook(input_file, solr_host="localhost:8983",test_only=False, verbose=False):
    print "Processing", input_file
    wb2 = openpyxl.load_workbook(input_file)

    prefix = os.path.basename(input_file)[:4]

    ws_plumes = wb2['%s_Plumes_IME'%prefix]
    ws_sources = wb2['%s_Sources'%prefix]

    plumes = read_plume_sheet(ws_plumes)
    sources = read_sources_sheet(ws_sources)

    upload_plumes_to_solr(plumes, solr_host=solr_host, test_only=test_only, verbose=verbose)
    upload_sources_to_solr(sources, solr_host=solr_host, test_only=test_only, verbose=verbose)

if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument("-d", "--data", help="Input Sources Lists (xlsx)", required=True, type=str, nargs='+')
    parser.add_argument("-s", "--solrhost", help="Target Solr host", type=str, default="localhost", required=False)
    parser.add_argument("-p", "--solrport", help="Target Solr port", type=str, default="8983", required=False)
    parser.add_argument("-v", "--verbose",
                        help="Extra output",
                        required=False, action="store_true")
    parser.add_argument("-t", "--test",
                        help="Test. Don't actually upload to Solr.",
                        required=False, action="store_true")


    args = parser.parse_args()
    input_files = args.data
    solr_host = SOLR_URL % (args.solrhost, args.solrport)
    verbose = args.verbose
    test_only = args.test

    for input_file in input_files:
        process_workbook(input_file, solr_host=solr_host, test_only=test_only, verbose=verbose)

    counts.commit_changes_to_solr(solr_host=solr_host, solr_core="MSF")